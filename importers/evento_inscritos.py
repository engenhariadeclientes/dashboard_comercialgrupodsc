"""Importer de listas de inscritos de eventos presenciais — spec seção 8.1.

Formato real confirmado em 22/07/2026 contra duas listas reais (Jornada Porter
Presidente Prudente e São Paulo): metadados no topo variam de layout entre eventos
(às vezes rótulo "Data:"/"Local:" numa linha e valor na linha de baixo, às vezes o
valor já vem direto sem rótulo), então a extração de metadados é best-effort por
regex/heurística — o que importa (telefone, nome, e-mail, dedupe, matching) é
robusto. Colunas além das mapeadas (Nº ingresso, Check-in, perguntas extras da
pesquisa) são preservadas em `payload`.

Uso: python importers/evento_inscritos.py arquivo1.xlsx [arquivo2.xlsx ...]
"""
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from importers.common_import import logger  # noqa: E402
from workers.common.db import ConexaoComReconexao  # noqa: E402
from workers.common.eventos import registrar_evento  # noqa: E402
from workers.common.matching import atualizar_campos_lead, resolver_ou_criar_lead  # noqa: E402
from workers.common.normalizacao import (  # noqa: E402
    derivar_marca,
    extrair_telefone_valido,
    gerar_variacoes_telefone,
    normalizar_email,
    normalizar_nome,
    normalizar_texto_busca,
)

FONTE = "importacao"
CANAL = "evento"

_RE_DATA = re.compile(r"(\d{2}/\d{2}/\d{4})(?:\s+(\d{1,2})h)?")

COLUNAS_MAPEADAS = {
    "ordem de inscrição", "nº ingresso", "nome", "sobrenome", "email", "telefone",
    "check-in", "cidade", "estado",
}


def _slug(texto: str) -> str:
    nfkd = unicodedata.normalize("NFKD", texto)
    sem_acento = "".join(c for c in nfkd if not unicodedata.combining(c))
    sem_acento = re.sub(r"[^A-Za-z0-9]+", "_", sem_acento).strip("_")
    return sem_acento


def _localizar_cabecalho(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    for i, row in enumerate(rows):
        valores = [str(c).strip() if c is not None else "" for c in row]
        if "Nome" in valores:
            mapa = {v.lower(): idx for idx, v in enumerate(valores) if v}
            return i, mapa
    raise ValueError("Linha de cabeçalho (com a coluna 'Nome') não encontrada")


def _extrair_metadados(rows_topo: list[tuple], nome_arquivo: str) -> dict:
    titulo = None
    data_evento = None
    local_evento = None

    for row in rows_topo:
        for cel in row:
            if not cel or not isinstance(cel, str):
                continue
            texto = cel.strip()
            if not texto or texto.lower() in ("data:", "local:"):
                continue
            m = _RE_DATA.search(texto)
            if m and data_evento is None:
                dia, hora = m.group(1), m.group(2)
                try:
                    data_evento = datetime.strptime(dia, "%d/%m/%Y")
                    if hora:
                        data_evento = data_evento.replace(hour=int(hora))
                except ValueError:
                    pass
                continue
            if titulo is None:
                titulo = texto
            elif local_evento is None and texto != titulo:
                local_evento = texto

    if titulo is None:
        titulo = Path(nome_arquivo).stem

    # campanha no padrão sugerido pela spec: Jornada_Porter_<Local>_<MesAno>
    # o texto após o "|" no título (ex.: "[2026] JORNADA PORTER | SÃO PAULO") é mais
    # confiável que o texto solto nos metadados (que às vezes é o endereço completo
    # do venue, não a cidade) — por isso tem prioridade quando existir.
    if "|" in titulo:
        nome_evento, local_part = [p.strip() for p in titulo.split("|", 1)]
        local_evento = local_part or local_evento
    else:
        nome_evento = titulo

    mes_ano = data_evento.strftime("%b%y").capitalize() if data_evento else ""
    campanha = _slug(f"{nome_evento}_{local_evento or ''}_{mes_ano}").strip("_")

    return {
        "campanha": campanha,
        "data_evento": data_evento or datetime.now(),
        "origem_detalhe": local_evento or titulo,
    }


def _valor(row: tuple, mapa: dict[str, int], chave: str):
    idx = mapa.get(chave)
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    return v.strip() if isinstance(v, str) else v


def importar_arquivo(conexao, caminho: str) -> tuple[int, int]:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    contadores = {"novos": 0, "existentes": 0}
    telefones_vistos_no_arquivo: set[str] = set()

    for nome_sheet in wb.sheetnames:
        ws = wb[nome_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        try:
            idx_cabecalho, mapa = _localizar_cabecalho(rows)
        except ValueError:
            logger.warning("Sheet '%s' de %s sem coluna 'Nome', ignorada", nome_sheet, caminho)
            continue

        meta = _extrair_metadados(rows[:idx_cabecalho], caminho)
        logger.info("Evento: campanha=%s data=%s local=%s", meta["campanha"], meta["data_evento"], meta["origem_detalhe"])

        for row in rows[idx_cabecalho + 1:]:
            if not any(row):
                continue

            nome_raw = " ".join(
                filter(None, [str(_valor(row, mapa, "nome") or ""), str(_valor(row, mapa, "sobrenome") or "")])
            ).strip()
            telefone_raw = _valor(row, mapa, "telefone")
            email_raw = _valor(row, mapa, "email")
            cidade = _valor(row, mapa, "cidade")
            uf_informada = _valor(row, mapa, "estado")
            checkin = _valor(row, mapa, "check-in")

            telefone, telefone_invalido = extrair_telefone_valido(str(telefone_raw)) if telefone_raw else (None, True)

            # dedupe obrigatório por telefone (spec 8.1) — mesma pessoa pode ter 2+ ingressos
            if telefone and telefone in telefones_vistos_no_arquivo:
                continue
            if telefone:
                telefones_vistos_no_arquivo.add(telefone)

            if not telefone and not email_raw and not nome_raw:
                continue

            nome = normalizar_nome(nome_raw)
            email = normalizar_email(email_raw)
            variacoes = gerar_variacoes_telefone(telefone) if telefone else []

            colunas_extra = {
                k: row[idx] for k, idx in mapa.items()
                if k not in COLUNAS_MAPEADAS and idx < len(row) and row[idx] is not None
            }
            payload = {"colunas_extra": colunas_extra, "arquivo": Path(caminho).name}
            if telefone_invalido and telefone_raw:
                payload["telefone_invalido"] = True

            def processar(conn, nome=nome, telefone=telefone, variacoes=variacoes, email=email, cidade=cidade,
                          uf_informada=uf_informada, payload=payload, checkin=checkin):
                marca_info = derivar_marca(conn, cidade=cidade, uf_informada=uf_informada, telefone_e164=telefone)
                dados = {
                    "nome": nome,
                    "telefone": telefone,
                    "telefone_variacoes": variacoes or None,
                    "email": email,
                    "cidade": cidade,
                    "uf": marca_info["uf"],
                    "regiao": marca_info["regiao"],
                    "marca": marca_info["marca"],
                    "uf_derivada_por_ddd": marca_info["uf_derivada_por_ddd"],
                    "canal_entrada": CANAL,
                    "campanha_entrada": meta["campanha"],
                    "origem_detalhe_entrada": meta["origem_detalhe"],
                    "tipo_captura": None,
                    "data_entrada": meta["data_evento"],
                    "payload": payload,
                }

                lead_id, criado = resolver_ou_criar_lead(conn, dados, fonte=FONTE)
                if not criado:
                    atualizar_campos_lead(
                        conn, lead_id,
                        {k: dados[k] for k in ("cidade", "uf", "regiao", "marca", "uf_derivada_por_ddd") if dados.get(k) is not None},
                    )
                    contadores["existentes"] += 1
                else:
                    contadores["novos"] += 1

                registrar_evento(conn, lead_id, FONTE, "importado", meta["data_evento"],
                                  canal=CANAL, campanha=meta["campanha"], origem_detalhe=meta["origem_detalhe"],
                                  payload=payload)
                if checkin and str(checkin).strip().lower() in ("sim", "yes", "true", "1"):
                    registrar_evento(conn, lead_id, FONTE, "checkin_evento", meta["data_evento"],
                                      canal=CANAL, campanha=meta["campanha"], origem_detalhe=meta["origem_detalhe"])
                conn.commit()

            conexao.executar(processar)
    wb.close()
    return contadores["novos"], contadores["existentes"]


def importar(caminhos: list[str]) -> None:
    total_novos, total_existentes = 0, 0
    conexao = ConexaoComReconexao()
    try:
        for caminho in caminhos:
            logger.info("Lendo %s", caminho)
            novos, existentes = importar_arquivo(conexao, caminho)
            total_novos += novos
            total_existentes += existentes
    finally:
        conexao.close()
    logger.info("Importação de eventos concluída: %d leads novos, %d já existentes (matched)", total_novos, total_existentes)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python importers/evento_inscritos.py arquivo1.xlsx [arquivo2.xlsx ...]", file=sys.stderr)
        sys.exit(1)
    importar(sys.argv[1:])
